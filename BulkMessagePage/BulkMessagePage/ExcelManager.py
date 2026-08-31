from openpyxl import load_workbook

#Reads contact data out of an uploaded Excel file
class ExcelManager:
    # Opens the Excel file and turns each row into a contract dictionary
    def read_contacts(self, file_path):
        workbook = load_workbook(file_path)
        sheet = workbook.active

        headers = []
        contacts = []

        # Read the column headers from row 1
        for cell in sheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip())

        # The file must have these two columns at minimum
        if "Name" not in headers or "Phone Number" not in headers:
            raise ValueError("Excel file must contain 'Name' and 'Phone Number' columns.")

        name_index = headers.index("Name")
        phone_index = headers.index("Phone Number")

        # Go through every row after the header row
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not row or all(value is None for value in row):
                continue # Skips empty rows

            name = row[name_index] if name_index < len(row) else None
            phone_number = row[phone_index] if phone_index < len(row) else None

            if not name or not phone_number:
                continue #Skips rows missing required infomation

            variable_data = {}

            # Store every column as extra data
            for index, header in enumerate(headers):
                if index < len(row):
                    variable_data[header] = row[index]

            contact = {
                "name": str(name).strip(),
                "phone_number": str(phone_number).strip(),
                "variable_data": variable_data
            }

            contacts.append(contact)

        return contacts
